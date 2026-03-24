# app/cli/stats.py
import asyncio
import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure project root in path
sys.path.append(os.getcwd())

from app.core.domain.historical_data import GAMES_1_7, PLAYER_PROPS, FAST_SET, POS_GRP
from app.db.database import async_session_maker
from app.db.models import Game, GameStats, User, Signup

# Excel Styles
HF = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HFont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
GOLD = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
GRN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAST = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GK_F = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DEF_F = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MID_F = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FWD_F = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
SECT = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SECT_FNT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
B = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
GK_FNT = Font(color="B7950B", bold=True)
DEF_FNT = Font(color="2471A3", bold=True)
MID_FNT = Font(color="1E8449", bold=True)
FWD_FNT = Font(color="CB4335", bold=True)

GRP_FILLS = {"ВРТ": GK_F, "ЗАЩ": DEF_F, "МИД": MID_F, "АТК": FWD_F}
GRP_FONT = {"ВРТ": GK_FNT, "ЗАЩ": DEF_FNT, "МИД": MID_FNT, "АТК": FWD_FNT}

def calc_win_pct(w, d, l):
    total = w + l + d
    if total == 0: return "-"
    return f"{w/total*100:.0f}%"

async def generate_stats_command(max_game_id: int, output_file: str = "FM_Player_Stats.xlsx"):
    """Consolidated command to generate Excel stats and Draft report."""
    print(f"Generating stats up to Game #{max_game_id}...")
    
    # 1. Initialize stats from Historical Data
    stats = {}
    for p in PLAYER_PROPS:
        stats[p] = {"elo": 100, "gp": 0, "g": 0, "w": 0, "l": 0, "d": 0, "mvp": 0, "logs": [""] * max_game_id}

    # 2. Process Historical Games 1-7 (or up to max_game_id if < 7)
    for g in GAMES_1_7:
        if g["id"] > max_game_id: break
        idx = g["id"] - 1
        w_team = g.get("W", [])
        l_team = g.get("L", [])
        d_team = g.get("D", [])
        goals = g.get("goals", {})
        
        for p in w_team:
            if p not in stats: stats[p] = {"elo": 100, "gp": 0, "g": 0, "w": 0, "l": 0, "d": 0, "mvp": 0, "logs": [""] * max_game_id}
            stats[p]["gp"] += 1; stats[p]["w"] += 1; stats[p]["elo"] += 10
            stats[p]["g"] += goals.get(p, 0)
            stats[p]["logs"][idx] = f"{goals.get(p,0)}g W"
            
        for p in l_team:
            if p not in stats: stats[p] = {"elo": 100, "gp": 0, "g": 0, "w": 0, "l": 0, "d": 0, "mvp": 0, "logs": [""] * max_game_id}
            stats[p]["gp"] += 1; stats[p]["l"] += 1
            # Special case for Game 1 draw logic preserved for historical accuracy
            if g["id"] == 1 and p in ["Arsen Kudaibergen", "Aidyn Galymbekov", "Umid Kamilov", "Daryn Bakyrkhan", "Zhanibek Xenbek", "Aktan Abdullaatov"]:
                stats[p]["d"] += 1; stats[p]["l"] -= 1; stats[p]["elo"] += 0
                stats[p]["g"] += goals.get(p, 0); stats[p]["logs"][idx] = f"{goals.get(p,0)}g D"
            else:
                stats[p]["elo"] -= 5; stats[p]["g"] += goals.get(p, 0); stats[p]["logs"][idx] = f"{goals.get(p,0)}g L"
                
        for p in d_team:
            if p not in stats: stats[p] = {"elo": 100, "gp": 0, "g": 0, "w": 0, "l": 0, "d": 0, "mvp": 0, "logs": [""] * max_game_id}
            stats[p]["gp"] += 1; stats[p]["d"] += 1; stats[p]["elo"] += 0
            stats[p]["g"] += goals.get(p, 0); stats[p]["logs"][idx] = f"{goals.get(p,0)}g D"

        for p in g.get("mvp", []):
            if p in stats: stats[p]["mvp"] += 1; stats[p]["elo"] += 5

    # 3. TODO: Process Games > 7 from Database if needed
    # For now, we replicate exactly what the original scripts did.

    # 4. Excel Generation
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Player Stats {max_game_id}G"

    headers = ["#", "Группа", "Игрок", "Поз", "Альт. позиции", "ELO", "Игры", "Голы", "Победы", "Пораж.", "Win%", "MVP", "Г/И", "Ораза"]
    for i in range(1, max_game_id + 1):
        headers.append(f"G{i}")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HFont; c.fill = HF; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = B
    
    # Sort and Section
    all_players = sorted(stats.keys(), key=lambda p: (-stats[p]["elo"], -stats[p]["g"], -stats[p]["gp"]))
    SECTIONS = [("ВРАТАРИ", []), ("ЗАЩИТА", []), ("ПОЛУЗАЩИТА", []), ("НАПАДЕНИЕ", []), ("БЕЗ СТАТЫ", [])]
    
    for p in all_players:
        st = stats[p]
        props = PLAYER_PROPS.get(p, {"pos": "CM", "alt": ""})
        grp = POS_GRP.get(props["pos"], "МИД")
        gpi = f"{st['g']/st['gp']:.1f}" if st['gp'] > 0 else "-"
        winpct = calc_win_pct(st['w'], st['d'], st['l'])
        d_tuple = (p, props["pos"], props.get("alt", ""), grp, st["elo"], st["gp"], st["g"], st["w"], st["l"], winpct, st["mvp"], gpi, "Да" if p in FAST_SET else "", st["logs"])
        
        if st["gp"] == 0: SECTIONS[4][1].append(d_tuple)
        elif grp == "ВРТ": SECTIONS[0][1].append(d_tuple)
        elif grp == "ЗАЩ": SECTIONS[1][1].append(d_tuple)
        elif grp == "МИД": SECTIONS[2][1].append(d_tuple)
        elif grp == "АТК": SECTIONS[3][1].append(d_tuple)

    row_idx = 2; num = 1
    for sect_name, players in SECTIONS:
        if not players: continue
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col); c.fill = SECT; c.border = B
        ws.cell(row=row_idx, column=2, value=sect_name).font = SECT_FNT
        row_idx += 1
        for d in players:
            vals = [num, d[3], d[0], d[1], d[2], d[4], d[5], d[6], d[7], d[8], d[9], d[10], d[11], d[12]] + d[13]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=col, value=val); c.border = B; c.alignment = Alignment(horizontal="center")
                if col == 3 or col == 5: c.alignment = Alignment(horizontal="left")
            
            if d[3] in GRP_FILLS:
                ws.cell(row=row_idx, column=2).fill = GRP_FILLS[d[3]]
                ws.cell(row=row_idx, column=4).fill = GRP_FILLS[d[3]]
            
            if d[10]: ws.cell(row=row_idx, column=12).fill = GOLD
            if d[12]: ws.cell(row=row_idx, column=14).fill = FAST
            
            # Color Game Logs
            for gcol in range(15, len(headers) + 1):
                c = ws.cell(row=row_idx, column=gcol); v = str(c.value or "")
                if "W" in v: c.fill = GRN
                elif "L" in v: c.fill = RED
                elif "D" in v: c.fill = GOLD
            
            row_idx += 1; num += 1

    wb.save(output_file)
    print(f"Excel saved to {output_file}")

async def generate_draft_report_command(date_str: str, output_file: str = "Draft_Stats.md"):
    """Placeholder or simplified draft report generator."""
    # Logic similar to generate_draft.py but integrated
    print(f"Generating Draft Report for {date_str}...")
    # ... logic omitted for brevity, could be added if needed ...
    pass
