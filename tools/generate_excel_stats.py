#!/usr/bin/env python3
"""Generate formatted XLSX grouped by position with corrected Game 6 roster positions."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HF = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HFont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
GOLD = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
GRN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAST = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GK_F = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DEF_F = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MID_F = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FWD_F = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
SECT = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SECT_FNT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
B = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
GK_FNT = Font(color="B7950B", bold=True)
DEF_FNT = Font(color="2471A3", bold=True)
MID_FNT = Font(color="1E8449", bold=True)
FWD_FNT = Font(color="CB4335", bold=True)

GRP_FILLS = {"ВРТ":GK_F,"ЗАЩ":DEF_F,"МИД":MID_F,"АТК":FWD_F}
POS_GRP = {"GK":"ВРТ","CB":"ЗАЩ","LB":"ЗАЩ","RB":"ЗАЩ",
           "CM":"МИД","CDM":"МИД","CAM":"МИД","LM":"МИД","RM":"МИД",
           "FWD":"АТК","LW":"АТК","RW":"АТК"}
GRP_FONT = {"ВРТ":GK_FNT,"ЗАЩ":DEF_FNT,"МИД":MID_FNT,"АТК":FWD_FNT}

# (name, pos, alt, grp, elo, gp, goals, w, l, wpct, mvp, gpi, fast, g1..g5)
SECTIONS = [
 ("ВРАТАРИ",[
  ("Dastan Sadyraliyev","GK","","ВРТ",110,4,0,2,2,"50%","",0.0,"","","0g W","0g L","0g W","0g L"),
  ("Aidyn Kassym","GK","","ВРТ",100,0,0,0,0,"-","","","","","","","",""),
 ]),
 ("ЗАЩИТА",[
  ("Asset Khassan","CB","","ЗАЩ",120,2,1,2,0,"100%","",0.5,"Да","","","1g W","","0g W"),
  ("Yernar Bestikpayev","LB","RB, CB, RW, LW","ЗАЩ",110,1,0,1,0,"100%","",0.0,"Да","","","","0g W",""),
  ("Iosif Vassilenko","RB","RW","ЗАЩ",110,1,0,1,0,"100%","",0.0,"","","","","","0g W"),
  ("Bakhtiyar Temirbayev","RB","RM, RW","ЗАЩ",105,2,3,1,1,"50%","",1.5,"Да","1g L","","","2g W",""),
  ("Akhmetov Dias","RB","RM","ЗАЩ",105,2,0,1,1,"50%","",0.0,"Да","","","","0g L","0g W"),
  ("Aktan Abdullaatov","RB","GK, LB, LW","ЗАЩ",105,3,0,1,1,"33%","",0.0,"","0g D","0g W","","0g L",""),
  ("Amir Rustam","RB","RW, CM","ЗАЩ",100,0,0,0,0,"-","",0.0,"Да","","","","",""),
  ("Almas Kaneshev","CB","CM","ЗАЩ",100,0,0,0,0,"-","",0.0,"","","","","",""),
  ("Niyaz Bayankulov","CB","","ЗАЩ",95,1,0,0,1,"0%","",0.0,"Да","","","","","0g L"),
  ("Medeu Yersain","CB","RB","ЗАЩ",95,1,0,0,1,"0%","",0.0,"Да","","","","0g L",""),
  ("Adil Baizhanov","CB","LB, RB, RW, LW","ЗАЩ",95,1,0,0,1,"0%","",0.0,"","","0g L","","",""),
  ("Yermakhan Borankhan","RB","","ЗАЩ",95,4,0,1,3,"25%","",0.0,"Да","","0g L","0g W","0g L","0g L"),
 ]),
 ("ПОЛУЗАЩИТА",[
  ("Said Soltanzade","CM","LW, CB, RB, LB, FWD, RW","МИД",135,5,4,4,1,"80%","",0.8,"","1g W","1g W","0g L","1g W","1g W"),
  ("Dias Bekbosyn","CAM","CM, FWD, CDM, LW, RW","МИД",130,4,5,3,1,"75%",1,1.2,"Да","5g W","0g L","0g W","","0g W"),
  ("Daryn Bakyrkhan","CM","CB, LB, RW","МИД",110,5,2,2,2,"40%","",0.4,"Да","0g D","0g W","1g L","0g W","1g L"),
  ("Iliyas Akbergen","CM","LM, RM, RW, LW","МИД",110,2,8,1,1,"50%",1,4.0,"","","2g L","6g W","",""),
  ("Aldiyar Kazbekov","CAM","CM, CDM","МИД",105,2,4,1,1,"50%","",2.0,"","2g W","","","","2g L"),
  ("Rauan Tanov","CM","RM, LM, LW","МИД",100,0,0,0,0,"-","","","","","","","",""),
  ("Adok Salimzhan","CM","","МИД",100,1,0,0,1,"0%",1,0.0,"Да","","","","0g L",""),
  ("Yerik Yekibay","CDM","","МИД",95,4,0,1,3,"25%","",0.0,"","","0g L","0g L","0g L","0g W"),
  ("Ansar Tanirbay","CDM","","МИД",95,1,0,0,1,"0%","",0.0,"Да","0g L","","","",""),
 ]),
 ("НАПАДЕНИЕ",[
  ("Vlad Lishanlo","LW","LW, RW, CM, LM, RM, LB, RB","АТК",120,5,15,3,2,"60%","",3.0,"","3g W","3g L","3g L","4g W","2g W"),
  ("Dias Kaibar","RW","LW","АТК",115,1,4,1,0,"100%",1,4.0,"","","","","","4g W"),
  ("Viktor Snytkin","FWD","FWD","АТК",115,3,5,2,1,"67%","",1.7,"","","","1g W","0g L","4g W"),
  ("Nurkabyl Sharipbay","FWD","CAM, RW, LW","АТК",115,1,7,1,0,"100%",1,7.0,"Да","","7g W","","",""),
  ("Miras Mukanov","RW","LW, CM, FWD","АТК",110,2,5,1,1,"50%",1,2.5,"","2g W","","","","3g L"),
  ("Assylbek Nurmagambetov","FWD","FWD, CAM","АТК",110,4,2,2,2,"50%","",0.5,"","0g L","0g W","0g L","","2g W"),
  ("Dias August","LW","RW, FWD","АТК",110,1,2,1,0,"100%","",2.0,"Да","","","","2g W",""),
  ("Mirlan Toktoshev","RW","LW, CM","АТК",100,3,3,1,2,"33%","",1.0,"Да","","2g W","","1g L","0g L"),
  ("Mukhammed A Askar","FWD","LB, CDM, CM","АТК",100,0,0,0,0,"-","","","Да","","","","",""),
  ("Mansur Mutayev","FWD","RW","АТК",90,2,5,0,2,"0%","",2.5,"","","","1g L","4g L",""),
 ]),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Player Stats"

headers = ["#","Группа","Игрок","Поз","Альт. позиции","ELO","Игры","Голы","Победы","Пораж.","Win%","MVP","Г/И","Ораза",
           "Игра 1\n17.01","Игра 2\n24.01","Игра 3\n31.01","Игра 4\n07.02","Игра 5\n14.02"]
for col,h in enumerate(headers,1):
    c = ws.cell(row=1,column=col,value=h)
    c.font = HFont; c.fill = HF
    c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border = B
widths = [4,8,26,5,26,6,6,6,8,8,7,5,6,7,10,10,10,10,10]
for i,w in enumerate(widths,1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 35

row = 2; num = 1
for sect_name, players in SECTIONS:
    for col in range(1,20):
        c = ws.cell(row=row,column=col); c.fill = SECT; c.border = B
    ws.cell(row=row,column=2,value=sect_name).font = SECT_FNT
    ws.cell(row=row,column=2).fill = SECT
    ws.cell(row=row,column=2).alignment = Alignment(horizontal="left")
    row += 1
    for d in players:
        vals = [num,d[3],d[0],d[1],d[2],d[4],d[5],d[6],d[7],d[8],d[9],d[10],d[11],d[12],d[13],d[14],d[15],d[16],d[17]]
        for col,val in enumerate(vals,1):
            c = ws.cell(row=row,column=col,value=val); c.border = B
            c.alignment = Alignment(horizontal="center")
            if col == 3: c.alignment = Alignment(horizontal="left")
            if col == 5: c.alignment = Alignment(horizontal="left")
        grp = d[3]
        if grp in GRP_FILLS:
            ws.cell(row=row,column=2).fill = GRP_FILLS[grp]
            ws.cell(row=row,column=4).fill = GRP_FILLS[grp]
        pos = d[1]
        pg = POS_GRP.get(pos,"")
        if pg in GRP_FONT: ws.cell(row=row,column=4).font = GRP_FONT[pg]
        # Alt pos font color by dominant group
        alt = d[2]
        if alt:
            al = [x.strip() for x in alt.split(",")]
            cnt = {"ЗАЩ":0,"МИД":0,"АТК":0,"ВРТ":0}
            for a in al:
                g = POS_GRP.get(a,"")
                if g: cnt[g] += 1
            dom = max(cnt, key=cnt.get)
            if cnt[dom] > 0 and dom in GRP_FONT:
                ws.cell(row=row,column=5).font = GRP_FONT[dom]
        if d[10]: ws.cell(row=row,column=12).fill = GOLD
        if d[12]: ws.cell(row=row,column=14).fill = FAST
        for gcol in range(15,20):
            c = ws.cell(row=row,column=gcol)
            v = str(c.value or "")
            if "W" in v: c.fill = GRN
            elif "L" in v: c.fill = RED
        row += 1; num += 1

ws.freeze_panes = "C2"

ws2 = wb.create_sheet("Games")
for col,h in enumerate(["Игра","Дата","Счёт","MVP"],1):
    c = ws2.cell(row=1,column=col,value=h); c.font = HFont; c.fill = HF; c.border = B
for i,g in enumerate([(1,"17.01","13:7:4","Dias Bekbosyn"),(2,"24.01","13:12","Nurkabyl Sharipbay"),
    (3,"31.01","8:11","Iliyas Akbergen"),(4,"07.02","10:5","Zhanibek Xenbek, Adok Salimzhan"),
    (5,"14.02","10:13","Dias Kaibar, Miras Mukanov")],2):
    for col,val in enumerate(g,1): ws2.cell(row=i,column=col,value=val).border = B

wb.save("/home/yeronym/Documents/fmBot/football-manager/FM_Player_Stats.xlsx")
