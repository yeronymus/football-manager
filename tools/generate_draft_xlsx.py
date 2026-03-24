#!/usr/bin/env python3
"""Generate formatted Draft XLSX for 28 players with G6 Draw logic."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HF = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HFont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
GOLD = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
GRN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FAST = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GK_F = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
DEF_F = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
MID_F = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FWD_F = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
SECT = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
SECT_FNT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
B = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
GK_FNT = Font(color="B7950B", bold=True)
DEF_FNT = Font(color="2471A3", bold=True)
MID_FNT = Font(color="1E8449", bold=True)
FWD_FNT = Font(color="CB4335", bold=True)

GRP_FILLS = {"ВРТ":GK_F,"ЗАЩ":DEF_F,"МИД":MID_F,"АТК":FWD_F}
POS_GRP = {"GK":"ВРТ","CB":"ЗАЩ","LB":"ЗАЩ","RB":"ЗАЩ",
           "CM":"МИД","CDM":"МИД","CAM":"МИД","LM":"МИД","RM":"МИД",
           "FWD":"АТК","LW":"АТК","RW":"АТК","ST":"АТК"}
GRP_FONT = {"ВРТ":GK_FNT,"ЗАЩ":DEF_FNT,"МИД":MID_FNT,"АТК":FWD_FNT}

games_raw = [
    {
        "id": 1, "score": "13:7:4", "W": ["Dias Bekbosyn","Vlad Lishanlo","Aldiyar Kazbekov","Miras Mukanov","Said Soltanzade","Yernur Bauyrzhanuly"],
        "L": ["Arsen Kudaibergen","Aidyn Galymbekov","Umid Kamilov","Daryn Bakyrkhan","Zhanibek Xenbek","Aktan Abdullaatov","Alikhan Muratuly","Yeldos Ismailov","Assylbek Nurmagambetov","Ansar Tanirbay","Bakhtiyar Temirbayev","Raimbek Sarzhakov"],
        "goals": {"Dias Bekbosyn":5,"Vlad Lishanlo":3,"Aldiyar Kazbekov":2,"Miras Mukanov":2,"Said Soltanzade":1,"Yernur Bauyrzhanuly":1,"Arsen Kudaibergen":3,"Aidyn Galymbekov":2,"Umid Kamilov":1,"Zhanibek Xenbek":1,"Yeldos Ismailov":2,"Bakhtiyar Temirbayev":1,"Raimbek Sarzhakov":1},
        "mvp": ["Dias Bekbosyn"]
    },
    {
        "id": 2, "score": "13:12", "W": ["Nurkabyl Sharipbay","Said Soltanzade","Mirlan Toktoshev","Daryn Bakyrkhan","Yernur Bauyrzhanuly","Aktan Abdullaatov","Zhanibek Xenbek","Dastan Sadyraliyev","Assylbek Nurmagambetov"],
        "L": ["Vlad Lishanlo","Iliyas Akbergen","Arsen Kudaibergen","Aidyn Galymbekov","Yeldos Ismailov","Yerik Yekibay","Dias Bekbosyn","Yermakhan Borankhan","Adil Baizhanov"],
        "goals": {"Nurkabyl Sharipbay":7,"Said Soltanzade":1,"Mirlan Toktoshev":2,"Vlad Lishanlo":3,"Iliyas Akbergen":2,"Arsen Kudaibergen":2,"Aidyn Galymbekov":2},
        "mvp": ["Nurkabyl Sharipbay"]
    },
    {
        "id": 3, "score": "8:11", "W": ["Beka","Daulet Bekmolda","Asset Khassan","Yermakhan Borankhan","Dias Bekbosyn","Iliyas Akbergen","Zhanibek Xenbek","Viktor Snytkin","Alikhan Muratuly"],
        "L": ["Dastan Sadyraliyev","Daryn Bakyrkhan","Vlad Lishanlo","Aidyn Galymbekov","Said Soltanzade","Yerik Yekibay","Assylbek Nurmagambetov","Mansur Mutayev","Raimbek Sarzhakov"],
        "goals": {"Asset Khassan":1,"Iliyas Akbergen":6,"Viktor Snytkin":1,"Daryn Bakyrkhan":1,"Vlad Lishanlo":3,"Aidyn Galymbekov":1,"Mansur Mutayev":1,"Raimbek Sarzhakov":1},
        "mvp": ["Iliyas Akbergen"]
    },
    {
        "id": 4, "score": "10:5", "W": ["Zhanibek Xenbek","Vlad Lishanlo","Daryn Bakyrkhan","Bakhtiyar Temirbayev","Yernar Bestikpayev","Said Soltanzade","Dias August","Dastan Sadyraliyev","Aidyn Galymbekov","Raimbek Sarzhakov"],
        "L": ["Adok Salimzhan","Akhmetov Dias","Mansur Mutayev","Medeu Yersain","Yerik Yekibay","Yermakhan Borankhan","Mirlan Toktoshev","Daulet Bekmolda","Viktor Snytkin","Aktan Abdullaatov"],
        "goals": {"Zhanibek Xenbek":1,"Vlad Lishanlo":4,"Bakhtiyar Temirbayev":2,"Said Soltanzade":1,"Dias August":2,"Mansur Mutayev":4,"Mirlan Toktoshev":1},
        "mvp": ["Zhanibek Xenbek", "Adok Salimzhan"]
    },
    {
        "id": 5, "score": "10:13", "W": ["Iosif Vassilenko","Akhmetov Dias","Asset Khassan","Dias Bekbosyn","Said Soltanzade","Yerik Yekibay","Dias Kaibar","Viktor Snytkin","Vlad Lishanlo","Assylbek Nurmagambetov"],
        "L": ["Dastan Sadyraliyev","Daulet Bekmolda","Yermakhan Borankhan","Niyaz Bayankulov","Aldiyar Kazbekov","Mirlan Toktoshev","Daryn Bakyrkhan","Aidyn Galymbekov","Raimbek Sarzhakov","Miras Mukanov"],
        "goals": {"Said Soltanzade":1,"Dias Kaibar":4,"Viktor Snytkin":4,"Vlad Lishanlo":2,"Assylbek Nurmagambetov":2,"Aldiyar Kazbekov":2,"Daryn Bakyrkhan":1,"Aidyn Galymbekov":2,"Raimbek Sarzhakov":1,"Miras Mukanov":3},
        "mvp": ["Dias Kaibar", "Miras Mukanov"]
    },
    {
        "id": 6, "score": "5:5", "W": [], "L": [],
        "D": ["Nurkabyl Sharipbay","Miras Mukanov","Iliyas Akbergen","Adok Salimzhan","Dias Bekbosyn","Mukhammed A Askar","Niyaz Bayankulov","Yermakhan Borankhan","Aidyn Kassym","Alikhan Muratuly", "Dias Kaibar","Daryn Bakyrkhan","Vlad Lishanlo","Ansar Tanirbay","Aldiyar Kazbekov","Mirlan Toktoshev","Asset Khassan","Almas Kaneshev","Yernar Bestikpayev","Aktan Abdullaatov"],
        "goals": {"Nurkabyl Sharipbay":4,"Miras Mukanov":1,"Vlad Lishanlo":1},
        "mvp": [] 
    }
]

stats = {}
for g in games_raw:
    idx = g["id"] - 1
    w_team = g.get("W", [])
    l_team = g.get("L", [])
    d_team = g.get("D", [])
    goals = g.get("goals", {})
    
    for p in w_team:
        if p not in stats: stats[p] = {"elo":100,"gp":0,"g":0,"w":0,"l":0,"d":0,"mvp":0,"logs":[""]*6}
        stats[p]["gp"] += 1; stats[p]["w"] += 1; stats[p]["elo"] += 10
        pgc = goals.get(p, 0)
        stats[p]["g"] += pgc
        stats[p]["logs"][idx] = f"{pgc}g W"
    for p in l_team:
        if p not in stats: stats[p] = {"elo":100,"gp":0,"g":0,"w":0,"l":0,"d":0,"mvp":0,"logs":[""]*6}
        stats[p]["gp"] += 1; stats[p]["l"] += 1
        if g["id"] == 1 and p in ["Arsen Kudaibergen","Aidyn Galymbekov","Umid Kamilov","Daryn Bakyrkhan","Zhanibek Xenbek","Aktan Abdullaatov"]:
            stats[p]["d"] += 1; stats[p]["l"] -= 1; stats[p]["elo"] += 0
            pgc = goals.get(p, 0)
            stats[p]["g"] += pgc
            stats[p]["logs"][idx] = f"{pgc}g D"
        else:
            stats[p]["elo"] -= 5
            pgc = goals.get(p, 0)
            stats[p]["g"] += pgc
            stats[p]["logs"][idx] = f"{pgc}g L"

    for p in d_team:
        if p not in stats: stats[p] = {"elo": 100, "gp":0, "g":0, "w":0, "l":0, "d":0, "mvp":0, "logs": [""]*6}
        stats[p]["gp"] += 1; stats[p]["d"] += 1; stats[p]["elo"] += 0
        pgc = goals.get(p, 0)
        stats[p]["g"] += pgc
        stats[p]["logs"][idx] = f"{pgc}g D"

    for p in g["mvp"]:
        if p in stats:
            stats[p]["mvp"] += 1; stats[p]["elo"] += 5

def calc_win_pct(w, d, l):
    total = w + l + d
    if total == 0: return "-"
    return f"{w/total*100:.0f}%"

FAST_SET = {
    "Yernar Bestikpayev","Akhmetov Dias","Bakhtiyar Temirbayev","Niyaz Bayankulov","Daulet Bekmolda",
    "Dias August","Adok Salimzhan","Ansar Tanirbay","Arsen Kudaibergen","Medeu Yersain","Yermakhan Borankhan",
    "Daryn Bakyrkhan","Mukhammed A Askar","Nurkabyl Sharipbay","Nurken Samigullin","Asset Khassan",
    "Mirlan Toktoshev","Yeldos Ismailov","Amir Rustam","Zhanibek Xenbek","Dias Bekbosyn"
}

DRAFT_28_DATA = [
    ("Yernar Bestikpayev", "RB", "CB, LB"),
    ("Yeldos Ismailov", "RW", "LW"),
    ("Adok Salimzhan", "CM", ""),
    ("Yernur Bauyrzhanuly", "LB", "RB"),
    ("Bakhtiyar Temirbayev", "LB", "LW"),
    ("Dias August", "RB", "RW, LW"),
    ("Dias Bekbosyn", "CM", ""),
    ("Mirlan Toktoshev", "CDM", "RW, LW"),
    ("Niyaz Bayankulov", "CB", "RB, LB"),
    ("Iliyas Akbergen", "CM", ""),
    ("Dastan Sadyraliyev", "GK", ""),
    ("Aidyn Galymbekov", "CM", "LB, RB, RW"),
    ("Nurkabyl Sharipbay", "ST", ""),
    ("Raimbek Sarzhakov", "FWD", "CAM, LW, RW, CM"),
    ("Nurken Samigullin", "RB", "LB, CM"),
    ("Vlad Lishanlo", "LB", "RB"),
    ("Aidyn Kassym", "GK", ""),
    ("Aldiyar Kazbekov", "CM", "CDM, CAM"),
    ("Miras Mukanov", "FWD", ""),
    ("Nurislam", "CM", "LB"),
    ("Zhanibek Xenbek", "LB", "RB, RW, LW"),
    ("Mansur Mutayev", "FWD", "LW, RW"),
    ("Assylbek Nurmagambetov", "CAM", "RW, LW"),
    ("Daryn Bakyrkhan", "CM", "CB, RW"),
    ("Yerik Yekibay", "CM", "LW"),
    ("Umid Kamilov", "CB", "RB, LB, ST"),
    ("Asset Khassan", "CB", "LB, RB"),
    ("Medeu Yersain", "CB", "RB, LB")
]

SECTIONS = {"ВРАТАРИ":[], "ЗАЩИТА":[], "ПОЛУЗАЩИТА":[], "НАПАДЕНИЕ":[]}
for name, mpos, altpos in DRAFT_28_DATA:
    st = stats.get(name, {"elo":100,"gp":0,"g":0,"w":0,"l":0,"d":0,"mvp":0,"logs":[""]*6})
    grp = POS_GRP.get(mpos.upper(), "МИД")
    winpct = calc_win_pct(st['w'], st['d'], st['l'])
    gpi = f"{st['g']/st['gp']:.1f}" if st['gp'] > 0 else "-"
    
    dtup = (name, mpos.upper(), altpos.upper(), grp, st["elo"], st["gp"], st["g"], st["w"], st["l"], winpct, st["mvp"], gpi, "Да" if name in FAST_SET else "", st["logs"][0], st["logs"][1], st["logs"][2], st["logs"][3], st["logs"][4], st["logs"][5])
    
    if grp == "ВРТ": SECTIONS["ВРАТАРИ"].append(dtup)
    elif grp == "ЗАЩ": SECTIONS["ЗАЩИТА"].append(dtup)
    elif grp == "МИД": SECTIONS["ПОЛУЗАЩИТА"].append(dtup)
    elif grp == "АТК": SECTIONS["НАПАДЕНИЕ"].append(dtup)

for k in SECTIONS: SECTIONS[k].sort(key=lambda x: (-x[4], -x[6], -x[5]))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Draft 28.02 List"

headers = ["#","Группа","Игрок","Поз","Альт. позиции","ELO","Игры","Голы","Победы","Пораж.","Win%","MVP","Г/И","Ораза",
           "G1\n17.01","G2\n24.01","G3\n31.01","G4\n07.02","G5\n14.02","G6\n22.02"]
for col,h in enumerate(headers,1):
    c = ws.cell(row=1,column=col,value=h)
    c.font = HFont; c.fill = HF
    c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border = B
widths = [4,8,26,5,26,6,6,6,8,8,7,5,6,7,9,9,9,9,9,9]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 35

row = 2; num = 1
for sect_name in ["ВРАТАРИ", "ЗАЩИТА", "ПОЛУЗАЩИТА", "НАПАДЕНИЕ"]:
    players = SECTIONS[sect_name]
    if not players: continue
    for col in range(1,21):
        c = ws.cell(row=row,column=col); c.fill = SECT; c.border = B
    ws.cell(row=row,column=2,value=sect_name).font = SECT_FNT
    ws.cell(row=row,column=2).fill = SECT
    ws.cell(row=row,column=2).alignment = Alignment(horizontal="left")
    row += 1
    for d in players:
        vals = [num,d[3],d[0],d[1],d[2],d[4],d[5],d[6],d[7],d[8],d[9],d[10],d[11],d[12],d[13],d[14],d[15],d[16],d[17],d[18]]
        for col,val in enumerate(vals,1):
            c = ws.cell(row=row,column=col,value=val); c.border = B
            c.alignment = Alignment(horizontal="center")
            if col == 3: c.alignment = Alignment(horizontal="left")
            if col == 5: c.alignment = Alignment(horizontal="left")
        grp = d[3]
        if grp in GRP_FILLS:
            ws.cell(row=row,column=2).fill = GRP_FILLS[grp]
            ws.cell(row=row,column=4).fill = GRP_FILLS[grp]
        pg = POS_GRP.get(d[1],"")
        if pg in GRP_FONT: ws.cell(row=row,column=4).font = GRP_FONT[pg]
        if d[10]: ws.cell(row=row,column=12).fill = GOLD
        if d[12]: ws.cell(row=row,column=14).fill = FAST
        for gcol in range(15,21):
            c = ws.cell(row=row,column=gcol)
            v = str(c.value or "")
            if "W" in v: c.fill = GRN
            elif "L" in v: c.fill = RED
            elif "D" in v: c.fill = GOLD
        row += 1; num += 1
ws.freeze_panes = "C2"
wb.save("/home/yeronym/Documents/fmBot/football-manager/FM_Draft_28_Stats.xlsx")
print("Saved FM_Draft_28_Stats.xlsx")
